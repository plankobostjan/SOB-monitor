import openpyxl
import time
import Tkinter as tk
import tkMessageBox
import pyttsx
import argparse
import sys
import os
import mysql.connector

class Bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def writeToSpreadsheet(sheet, happiness, energy, focus, spreadsheet_line_to_write):
    try:
        pos = "A" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. A2, from column name + spreadsheet_line_to_write (line number)
        sheet[pos] = time.strftime("%d.%m.%Y") #write date (dd.mm.yyyy) to sheet on spreadsheet_line_to_write eg. A2

        pos = "B" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. B2, from column name + spreadsheet_line_to_write (line number)
        sheet[pos] = time.strftime("%H:%M:%S") #write time (hh:mm:ss) to sheet on spreadsheet_line_to_write eg. B2

        pos = "C" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. C2, from column name + spreadsheet_line_to_write (line number)
        sheet[pos] = energy #write energy value to sheet on spreadsheet_line_to_write eg. C2

        pos = "D" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. D2, from column name + spreadsheet_line_to_write (line number)
        sheet[pos] = happiness #write happiness value to sheet on spreadsheet_line_to_write eg. D2

        pos = "E" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. E2, from column name + spreadsheet_line_to_write (line number)
        sheet[pos] = focus #write focus value to sheet on spreadsheet_line_to_write eg. E2

        spreadsheet_line_to_write += 1 #increase value of spreadsheet_line_to_write (sheet line) by one
        sheet['G2'] = spreadsheet_line_to_write #write spreadsheet_line_to_write value to the shell on spreadsheet_line_to_write G2

        print Bcolors.OKGREEN + "Data successfuly entered into a sheet!" + Bcolors.ENDC

    except:
        print Bcolors.FAIL + str(sys.exc_info()[0]) + Bcolors.ENDC
        print Bcolors.WARNING + "Something went wrong when entering data into a sheet!\nTry again!" + Bcolors.ENDC
        sys.exit()

def saveSpreadsheet(wb):
    try:
        wb.save("SOBm-data.xlsx") #save spreadsheet to /home/bostjan/...
        print Bcolors.OKGREEN + "File saved successfully!" + Bcolors.ENDC

    except:
        print Bcolors.WARNING + "Something went wrong while saving a file! File not saved!/nTry again!" + Bcolors.ENDC
        
def getAveragesFromSpreadsheet(worksheet):
 try:
     #start of part for getting data from sheet
     wb = openpyxl.load_workbook("SOBm-data.xlsx")
     sheet = wb.get_sheet_by_name(worksheet)
     average_happiness = 0
     average_energy = 0
     average_focus = 0
     for char in xrange(67, 70):
         avr = 0.0
         for row in xrange(2, sheet.max_row):
             if  sheet[str(chr(char)) + str(row)].value == None:
                 break
             #print sheet[str(chr(char)) + str(row)].value
             avr += sheet[str(chr(char)) + str(row)].value
         if char == 67:
             average_energy = avr / (sheet.max_row - 1)
         elif char == 68: 
             average_happiness = avr / (sheet.max_row -1)
         else:
             average_focus = avr / (sheet.max_row -1)
     print "Average happiness: {:.2f}\nAverage energy: {:.2f}\nAverage focus: {:.2f}".format(average_energy, average_happiness, average_focus)

 except IOError:
     print Bcolors.WARNING + "No such file or directory: {}\nProgram exited!".format(filename) + Bcolors.ENDC
     sys.exit()
 except KeyError:
     print Bcolors.WARNING + "Worksheet '{}' does not exist.\nProgram exited!".format(worksheet) + Bcolors.ENDC
     sys.exit()
 except:
     print Bcolors.WARNING + "Something went wrong!\nProgram exited!" + Bcolors.ENDC
     sys.exit()

def createSpreadsheet():
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "data"
        sheet["A1"] = "Date"
        sheet["B1"] = "Time"
        sheet["C1"] = "Energy"
        sheet["D1"] = "Happiness"
        sheet["E1"] = "Focus"
        sheet["G1"] = "Spreadsheet_Line_To_Write"
        sheet["G2"] = 2
        wb.save("SOBm-data.xlsx")
        print Bcolors.OKGREEN + "Spreadsheet successfully crated!" + Bcolors.ENDC
        
    except:
        print Bcolors.WARNING + "Something went wrong!\nSpreadsheet was not created!\nTry again!" + Bcolors.ENDC
    
def spreadsheetMode(values, sheetName):
    try:
        wb = openpyxl.load_workbook("SOBm-data.xlsx")
        sheet = wb.get_sheet_by_name(sheetName)
        spreadsheet_line_to_write = sheet['G2'].value
        writeToSpreadsheet(sheet, values[0], values[1], values[2], spreadsheet_line_to_write)
        saveSpreadsheet(wb)

    except IOError:
        print Bcolors.WARNING + "No such file or directory: {}\nProgram exited!".format("SOB-data.xslx") + Bcolors.ENDC
        sys.exit()
    except KeyError:
        print Bcolors.WARNING + "Worksheet '{}' does not exist.\nProgram exited!".format(sheetName) + Bcolors.ENDC
        sys.exit()
    except:
        print Bcolors.WARNING + "Something went wrong!\nProgram exited!" + Bcolors.ENDC
        sys.exit()

def getAveragesFromDatabase():
    try:
           #start of part for getting data from database
           con = mysql.connector.connect(user='', password='', host='',
                                     database='sob_monitor')
           cursor = con.cursor()
           get_number_of_rows = "SELECT COUNT(*) FROM sob_data;"
           get_happiness = "SELECT happiness FROM sob_data;"
           get_energy = "SELECT energy FROM sob_data;"
           get_focus = "SELECT focus FROM sob_data;"
           cursor.execute(get_number_of_rows)
           for value in cursor:
               number_of_rows = value[0]

           cursor.execute(get_happiness)
           happiness = 0
           for value in cursor:
               happiness += value[0]
           average_happiness = happiness / number_of_rows

           cursor.execute(get_energy)
           energy = 0
           for value in cursor:
               energy += value[0]
           average_energy = energy / number_of_rows

           cursor.execute(get_focus)
           focus = 0
           for value in cursor:
               focus += value[0]
           average_focus = focus / number_of_rows
           con.close()
           #end of part for getting data from server
           print "Average happiness: {:.2f}\nAverage energy: {:.2f}\nAverage focus: {:.2f}".format(average_energy, average_happiness, average_focus)

    except:
        print Bcolors.WARNING + "Something went wrong!\nProgram exited!" + Bcolors.ENDC
        print Bcolors.FAIL + sys.exc_info()[0] + Bcolors.ENDC
        sys.exit()

def writeToDatabase(date, hour, happiness, energy, focus):
    con = mysql.connector.connect(user='', password='', host='', database='sob_monitor')
    cursor = con.cursor()
    data = "INSERT INTO sob_data VALUES({}, {}, {}, {}, {}, {});".format(0, date, hour, happiness, energy, focus)
    cursor.execute(data)
    con.close()

def databaseMode(values):
    try:
        hour = time.strftime("%H%M%S")
        date = time.strftime("%Y%m%d")
        #print values
        writeToDatabase(date, hour, values[0], values[1], values[2])
        print Bcolors.OKGREEN + "Data was successfully written to database!" + Bcolors.ENDC

    except Exception, e:
        print Bcolors.WARNING + "Failed writing data to database!" + Bcolors.ENDC
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()

def restartProgram():
        python = sys.executable
        os.execl(python, python, * sys.argv)
    
def popupMessage(message): #show popup message
    root = tk.Tk()
    root.withdraw()
    tkMessageBox.showwarning('Self Stats', message)

def voiceMessage(text): #voice message
    engine = pyttsx.init()
    engine.say(text)
    engine.runAndWait()

def user_input(): #take user_input of happiness, focus and energy
    while True:
        
        try:
            happiness = int(raw_input("Enter happiness level between 1 and 10: ")) #happiness input
            if happiness < 1 or happiness > 10: #check if happiness is in range 1 to 10
                print Bcolors.WARNING + "Sorry, wrong input! Try again." + Bcolors.ENDC
                continue
            energy = int(raw_input("Enter energy level between 1 and 10: ")) #energy input
            if energy < 1 or energy > 10: #check if energy is in range 1 to 10
               print Bcolors.WARNING + "Sorry, wrong input! Try again." + Bcolors.ENDC
               continue
            focus = int(raw_input("Enter focus level between 1 and 10: ")) #focus input
            if focus < 1 or focus > 10: #check if focus is is range 1 to 10
               print Bcolors.WARNING + "Sorry, wrong input! Try again." + Bcolors.ENDC
               continue
        except:
            print Bcolors.WARNING + "Something went wrong! Try again!" + Bcolors.ENDC
            continue
        
        if 0 < focus <=10 and 0 < happiness <= 10 and 0 < energy <= 10: #when all values are between 1 and 10 brake the loop
            #print Bcolors.OKGREEN + "Input successful!" + Bcolors.ENDC
            return [happiness, energy, focus] #return values as a list
            break
        
def wait(): #defines time between inputs
    print time.strftime("%H:%M:%S")
    print "Waiting 1 hour."
    time.sleep(3600)

def arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("-c", "--create-spreadsheet", action="store_true",
                        help="Create new spreadsheet in current folder.")
    parser.add_argument('-a', '--show-worksheet-averages', action="store_true",
                        help="Show averages for selected spreadsheet.")
    parser.add_argument('--database-mode', action="store_true",
                        help="Use database (you need to create it) instead of spreadsheet")
    parser.add_argument("--sheet-name", type=str, help="Name of the sheet in wich data will be saved. Default: 'data'")
    args = parser.parse_args()

    if args.database_mode == True:
        global mode
        mode = "database"
    else:
        mode = "spreadsheet"

    if args.create_spreadsheet == True:
        createSpreadsheet()
        sys.exit()

    if args.show_worksheet_averages == True and mode == "database":
        arg = args.show_worksheet_averages
        getAveragesFromDatabase()
        sys.exit()

    elif args.show_worksheet_averages == True and mode == "spreadsheet":
        sheet = raw_input("Enter the name of sheet: ")
        getAveragesFromSpreadsheet(sheet)
        sys.exit()

    if args.sheet_name is not None:
        global sheet_name
        sheet_name = args.sheet_name
    else:
        sheet_name = "data"

def main():

    arguments()
    
    while True:
        voiceMessage("It's time to enter your feelings.")
        popupMessage("It's time to monitor your feelings")
        values = user_input()
        if mode == "spreadsheet":
            spreadsheetMode(values, sheet_name)
        else:
            databaseMode(values)
        wait()
        
if __name__=="__main__":
    main()
